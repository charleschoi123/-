import os, io, re, json, csv, time, uuid, threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import (
    Flask, request, render_template, render_template_string,
    send_file, Response, redirect, url_for, jsonify
)
from werkzeug.utils import secure_filename

import docx2txt, pdfplumber
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.chart import RadarChart, Reference
from openpyxl.chart.marker import DataPoint

# ===================== 基础配置（来自环境变量） =====================
DEEPSEEK_API_KEY   = os.getenv("DEEPSEEK_API_KEY", "")
OPENAI_BASE_URL    = os.getenv("OPENAI_BASE_URL", "https://api.deepseek.com")
LLM_MODEL_CHEAP    = os.getenv("LLM_MODEL_CHEAP", "deepseek-chat")

MAX_FILES          = int(os.getenv("MAX_FILES", "300"))
MAX_CONTENT_CHARS  = int(os.getenv("MAX_CONTENT_CHARS", "40000"))
MUST_HAVE_CAP      = int(os.getenv("MUST_HAVE_CAP", "60"))
MAX_WORKERS        = int(os.getenv("MAX_WORKERS", "3"))
STREAM_MODE        = os.getenv("STREAM_MODE", "sse")  # "sse" 或 "poll"

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB
app.secret_key = os.getenv("FLASK_SECRET", "dev_secret")

# OpenAI 兼容 DeepSeek
client = OpenAI(api_key=DEEPSEEK_API_KEY, base_url=OPENAI_BASE_URL)

# 批次状态（内存）
# SINGLE：{...,"report_lang": "zh|en|bi", rows:[], csv, excel}
# MULTI ：{...,"report_lang": "zh|en|bi", pairs:[], csv, excel}
BATCHES = {}

ALLOWED_EXT = {".pdf", ".docx", ".txt", ".doc"}
def allowed(filename: str) -> bool:
    return any(filename.lower().endswith(ext) for ext in ALLOWED_EXT)

# --------------------- 文本解析（含 PDF 兜底） ---------------------
def parse_text_from_blob(filename: str, blob: bytes) -> str:
    name = secure_filename(filename).lower()
    _, ext = os.path.splitext(name)
    content = ""
    try:
        if ext == ".pdf":
            try:
                with pdfplumber.open(io.BytesIO(blob)) as pdf:
                    pages = [(p.extract_text() or "") for p in pdf.pages]
                content = "\n".join(pages).strip()
            except Exception:
                content = ""
            if len(content) < 40:
                try:
                    import pypdfium2 as pdfium
                    pdf = pdfium.PdfDocument(io.BytesIO(blob))
                    texts = []
                    for i in range(len(pdf)):
                        page = pdf[i]
                        txtpage = page.get_textpage()
                        texts.append(txtpage.get_text_range())
                        txtpage.close()
                    content = "\n".join(texts).strip()
                except Exception:
                    pass
        elif ext == ".docx":
            try:
                content = docx2txt.process(io.BytesIO(blob)) or ""
            except Exception:
                content = ""
        elif ext in [".txt", ".doc"]:
            try:
                content = blob.decode("utf-8", errors="ignore")
            except Exception:
                content = blob.decode("latin-1", errors="ignore")
        else:
            content = ""
    except Exception:
        content = ""
    content = re.sub(r"\s+", " ", (content or "")).strip()
    if len(content) > MAX_CONTENT_CHARS:
        content = content[:MAX_CONTENT_CHARS] + "\n[TRUNCATED]"
    return content

# --------------------- MUST 抽取 + 年龄兜底 ---------------------
def extract_must_from_notes(notes: str):
    must = []
    for line in (notes or "").splitlines():
        s = line.strip()
        if not s:
            continue
        if re.match(r'^(!|\[?MUST\]?|必须|必需|必备)\s*[:：\-]?', s, flags=re.I):
            s = re.sub(r'^(!|\[?MUST\]?|必须|必需|必备)\s*[:：\-]?\s*', '', s, flags=re.I)
            if s:
                must.append(s)
    return must

UNDERGRAD_PAT = re.compile(r'(本科|学士|Bachelor)[^0-9]{0,12}(\d{4})(?:\D{0,3}(\d{4}))?', re.I)
def estimate_age_from_text(txt: str):
    m = UNDERGRAD_PAT.search(txt or "")
    if not m:
        return "不详"
    try:
        start_year = int(m.group(2))
        birth_year = start_year - 18
        return f"约{birth_year}年生"
    except Exception:
        return "不详"

# --------------------- 工具：JSON 容错解析 ---------------------
def _safe_json_parse(raw: str):
    import json as _json, re as _re
    if raw is None:
        raise ValueError("empty content")
    s = raw.strip()
    try:
        return _json.loads(s)
    except Exception:
        pass
    s2 = _re.sub(r"^```(?:json)?\s*|\s*```$", "", s, flags=_re.I | _re.S).strip()
    try:
        return _json.loads(s2)
    except Exception:
        pass
    i, j = s.find("{"), s.rfind("}")
    if i != -1 and j != -1 and j > i:
        return _json.loads(s[i:j+1])
    return _json.loads(s)

# --------------------- 语言映射 ---------------------
HEADERS = {
    "zh": ["文件名","姓名","年龄","教育背景","履历概要","亮点","匹配度","匹配度分析","证据","风险备注",
           "技能匹配","教育匹配","经验匹配","语言能力","稳定性"],
    "en": ["File","Name","Age","Education","Summary","Highlights","Overall","Fit Analysis","Evidence","Risk Notes",
           "Skills","EducationScore","Experience","Language","Stability"],
    "bi": ["文件名/File","姓名/Name","年龄/Age","教育/Education","概要/Summary","亮点/Highlights","总分/Overall",
           "分析/Analysis","证据/Evidence","风险/Risk",
           "技能/Skills","教育/EduScore","经验/Experience","语言/Language","稳定性/Stability"]
}
DIM_KEYS = ["skills","education","experience","language","stability"]

def label_lang(lang, zh, en):
    if lang == "en":
        return en
    if lang == "bi":
        return f"{zh}/{en}"
    return zh  # zh

# --------------------- 提示词与调用（含语言+维度评分） ---------------------
def build_prompt(jd_text: str, notes: str, must_list: list[str], resume_text: str, lang: str = "zh"):
    # 回复语言指示
    lang_line = {
        "zh": "Answer all fields in Chinese.",
        "en": "Answer all fields in English.",
        "bi": "Answer each field in Chinese first then English, separated by ' / '."
    }.get(lang, "Answer all fields in Chinese.")

    sys_prompt = (
        "You are a professional recruiter across industries. "
        "Evaluate the resume against the JD and notes. "
        f"If any MUST-HAVE is missing, overall score must be <= {MUST_HAVE_CAP}. "
        "Derive criteria from the JD; do not use canned ATS rubrics. "
        "Return STRICT JSON only. "
        "Estimate age ONLY from undergraduate enrollment year (assume 18). If unknown, return '不详'. "
        + lang_line
    )

    # 新增维度评分说明
    dims = """
Scoring dimensions (0-100 each):
- skills: core hard/soft skills relevant to JD
- education: degree and school alignment to JD
- experience: industry/domain/years/role relevance
- language: languages required by JD
- stability: timeline consistency, tenure, gaps (higher is better)
Overall can be the model's holistic judgment, but must be <= {cap} if any MUST-HAVE missing.
""".format(cap=MUST_HAVE_CAP)

    user_prompt = f"""
[JD]
{jd_text or '(none)'}

[NOTES]
{notes or '(none)'}

[MUST-HAVE]
- {(chr(10)+'- ').join(must_list) if must_list else '(none)'}

[RESUME]
{resume_text}

{dims}

[OUTPUT SCHEMA]
{{
  "name": "候选人姓名或未知 / Unknown if not found",
  "education_brief": "1-2行教育背景要点 / 1-2 lines education",
  "estimated_age": "约1989年生 或 不详 / ~1989 or Unknown",
  "summary": "2-3行履历概要 / 2-3 lines summary",
  "highlights": ["亮点1 / highlight1","亮点2 / highlight2"],
  "fit_analysis": "1-2段，结合JD说明匹配与不匹配点；如缺MUST要指出 / 1-2 paragraphs fit analysis; note missing MUSTs",
  "subscores": {{
      "skills": 0-100,
      "education": 0-100,
      "experience": 0-100,
      "language": 0-100,
      "stability": 0-100
  }},
  "overall": 0-100,
  "evidence": ["“原文摘录A … / quote A …”","“原文摘录B … / quote B …”"],
  "risk_notes": ["稳定性/时间线缺口等（如有） / risks if any"]
}}
Only return JSON.
    """.strip()
    return sys_prompt, user_prompt

def call_llm(sys_prompt: str, user_prompt: str, retries: int = 3):
    backoff = 0.8
    last_err = None
    raw_preview = ""
    for _ in range(retries):
        try:
            resp = client.chat.completions.create(
                model=LLM_MODEL_CHEAP,
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.1,
                max_tokens=1400,
                response_format={"type": "json_object"},
            )
            raw = resp.choices[0].message.content
            raw_preview = (raw or "")[:200]
            data = _safe_json_parse(raw or "")
            # 字段兜底
            data.setdefault("name", "未知")
            data.setdefault("education_brief", "")
            data.setdefault("estimated_age", "不详")
            data.setdefault("summary", "")
            data.setdefault("highlights", [])
            data.setdefault("fit_analysis", "")
            data.setdefault("overall", 0)
            data.setdefault("evidence", [])
            data.setdefault("risk_notes", [])
            ss = data.get("subscores") or {}
            data["subscores"] = {k:int(ss.get(k,0) or 0) for k in DIM_KEYS}
            return True, data
        except Exception as e:
            last_err = e
            time.sleep(backoff); backoff *= 1.6
    err_msg = f"{type(last_err).__name__}: {last_err}"
    if raw_preview:
        err_msg += f" | raw: {raw_preview}"
    return False, {"error": err_msg}

# --------------------- 单职位：逐简历处理 ---------------------
def process_one(blob_item: dict, jd_text: str, notes: str, must_list: list[str], report_lang: str):
    filename = blob_item["filename"]
    blob = blob_item["blob"]
    text = parse_text_from_blob(filename, blob)
    if not text:
        return {"filename": filename, "ok": False, "data": {
            "name":"未知","education_brief":"","estimated_age":"不详","summary":"",
            "highlights":[],"fit_analysis":"","overall":0,"evidence":[],
            "risk_notes":["解析失败或空文档"],"subscores":{k:0 for k in DIM_KEYS}
        }}
    sys_p, user_p = build_prompt(jd_text, notes, must_list, text, lang=report_lang)
    ok, data = call_llm(sys_p, user_p)
    if ok and (not data.get("estimated_age") or data["estimated_age"] == "不详"):
        data["estimated_age"] = estimate_age_from_text(text)
    if not ok:
        err = data.get("error","LLM失败")
        return {"filename": filename, "ok": False, "data": {
            "name":"未知","education_brief":"","estimated_age":"不详","summary":"",
            "highlights":[],"fit_analysis":"","overall":0,"evidence":[],
            "risk_notes":[err],"subscores":{k:0 for k in DIM_KEYS}
        }}
    return {"filename": filename, "ok": True, "data": data}

# --------------------- Excel 工具：插入雷达图 ---------------------
def add_radar_for_top(ws, start_row, start_col, top_rows):
    """
    在工作表 ws 上，从 start_row/start_col 起，基于 top_rows（[ [name, skills, edu, exp, lang, stab], ... ]）
    画一个雷达图。每一行一个系列。
    """
    if not top_rows:
        return
    # 写数据表头
    headers = ["Name","skills","education","experience","language","stability"]
    ws.cell(row=start_row, column=start_col, value=headers[0])
    for i, h in enumerate(headers[1:], start=1):
        ws.cell(row=start_row, column=start_col+i, value=h)
    # 写数据
    for r_i, row in enumerate(top_rows, start=1):
        for c_i, val in enumerate(row, start=0):
            ws.cell(row=start_row+r_i, column=start_col+c_i, value=val)

    # 构建雷达图
    chart = RadarChart()
    chart.type = "filled"
    # 类别（维度标签）
    cats = Reference(ws, min_col=start_col+1, min_row=start_row, max_col=start_col+5, max_row=start_row)
    # 为每个候选人添加系列
    for idx in range(len(top_rows)):
        data_ref = Reference(ws,
                             min_col=start_col+1, max_col=start_col+5,
                             min_row=start_row+1+idx, max_row=start_row+1+idx)
        series = chart.series.append(data_ref)
        chart.series[idx].title = ws.cell(row=start_row+1+idx, column=start_col).value
        chart.set_categories(cats)
    chart.height = 18
    chart.width = 24
    ws.add_chart(chart, ws.cell(row=start_row, column=start_col+7).coordinate)

# ===========================================================
# 路由 ———— 单职位模式（保留）
# ===========================================================
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")

@app.route("/start", methods=["POST"])
def start():
    jd_text = request.form.get("jd_raw", "").strip()
    notes = request.form.get("notes", "").strip()
    report_lang = (request.form.get("report_lang") or request.args.get("report_lang") or "zh").lower()
    if report_lang not in ("zh","en","bi"): report_lang = "zh"

    uploads = []
    for f in request.files.getlist("resumes"):
        if not f or not allowed(f.filename): continue
        uploads.append({"filename": f.filename, "blob": f.read()})
    if not uploads:
        return "No valid files. Allowed: .pdf, .docx, .txt, .doc", 400
    if len(uploads) > MAX_FILES:
        return f"Too many files (>{MAX_FILES}).", 400

    batch_id = str(uuid.uuid4())
    must_list = extract_must_from_notes(notes)
    BATCHES[batch_id] = {
        "mode": "single",
        "jd": jd_text,
        "notes": notes,
        "must": must_list,
        "report_lang": report_lang,
        "rows": [],
        "done": False,
        "csv": None,
        "excel": None,
    }

    def run_batch():
        rows = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futs = [ex.submit(process_one, item, jd_text, notes, must_list, report_lang) for item in uploads]
            for fu in as_completed(futs):
                r = fu.result(); rows.append(r); BATCHES[batch_id]["rows"].append(r)

        rows.sort(key=lambda x: x["data"].get("overall", 0), reverse=True)

        # CSV
        csv_buf = io.StringIO(); w = csv.writer(csv_buf)
        hdr = ["filename","name","estimated_age","overall","education_brief","summary","highlights","fit_analysis",
               "risk_notes"] + [f"sub_{k}" for k in DIM_KEYS]
        w.writerow(hdr)
        for r in rows:
            d = r["data"]; ss = d.get("subscores",{})
            w.writerow([
                r["filename"], d.get("name",""), d.get("estimated_age",""), d.get("overall",0),
                d.get("education_brief",""), d.get("summary",""),
                " | ".join(d.get("highlights",[])),
                (d.get("fit_analysis","") or "").replace("\n"," "),
                " | ".join(d.get("risk_notes",[])),
                ss.get("skills",0), ss.get("education",0), ss.get("experience",0), ss.get("language",0), ss.get("stability",0)
            ])
        BATCHES[batch_id]["csv"] = io.BytesIO(csv_buf.getvalue().encode("utf-8"))

        # Excel（含雷达图）
        lang = report_lang
        wb = Workbook(); ws = wb.active; ws.title = "Results"
        ws.append(HEADERS[lang])
        for r in rows:
            d = r["data"]; ss = d.get("subscores",{})
            ws.append([
                r["filename"], d.get("name",""), d.get("estimated_age",""),
                d.get("education_brief",""), d.get("summary",""),
                " | ".join(d.get("highlights",[])),
                d.get("overall",0), d.get("fit_analysis",""),
                " | ".join(d.get("evidence",[])), " | ".join(d.get("risk_notes",[])),
                ss.get("skills",0), ss.get("education",0), ss.get("experience",0), ss.get("language",0), ss.get("stability",0)
            ])

        # 雷达图（Top3）
        top3 = []
        for r in rows[:3]:
            d = r["data"]; ss = d.get("subscores",{})
            nm = d.get("name","") or r["filename"]
            top3.append([nm, ss.get("skills",0), ss.get("education",0), ss.get("experience",0), ss.get("language",0), ss.get("stability",0)])
        if top3:
            add_radar_for_top(ws, start_row=2+len(rows)+2, start_col=2, top_rows=top3)  # 放在表格下方

        excel_buf = io.BytesIO(); wb.save(excel_buf); excel_buf.seek(0)
        BATCHES[batch_id]["excel"] = excel_buf
        BATCHES[batch_id]["done"] = True

    threading.Thread(target=run_batch, daemon=True).start()
    return redirect(url_for("results", batch_id=batch_id))

@app.route("/results/<batch_id>")
def results(batch_id):
    b = BATCHES.get(batch_id)
    if not b or b.get("mode") != "single":
        return "Batch not found", 404
    return render_template(
        "results.html",
        batch_id=batch_id,
        jd=b["jd"], notes=b["notes"], must=b["must"],
        stream_mode=STREAM_MODE,
    )

@app.route("/events/<batch_id>")
def events(batch_id):
    def gen():
        sent = 0
        while True:
            b = BATCHES.get(batch_id)
            if not b: break
            rows = b["rows"]
            while sent < len(rows):
                yield f"data: {json.dumps(rows[sent], ensure_ascii=False)}\n\n"; sent += 1
            if b["done"]:
                yield "event: done\ndata: end\n\n"; break
            time.sleep(0.8)
    return Response(gen(), mimetype="text/event-stream")

@app.route("/rows/<batch_id>")
def rows_json(batch_id):
    b = BATCHES.get(batch_id)
    if not b: return jsonify({"error":"Batch not found"}), 404
    return jsonify({"rows": b.get("rows", []), "done": b["done"]})

@app.route("/download_csv/<batch_id>")
def download_csv(batch_id):
    b = BATCHES.get(batch_id)
    if not b or not b.get("csv"): return "CSV not ready", 404
    b["csv"].seek(0)
    return send_file(b["csv"], as_attachment=True,
                     download_name=f"results_{batch_id[:8]}.csv",
                     mimetype="text/csv")

@app.route("/download_excel/<batch_id>")
def download_excel(batch_id):
    b = BATCHES.get(batch_id)
    if not b or not b.get("excel"): return "Excel not ready", 404
    b["excel"].seek(0)
    return send_file(b["excel"], as_attachment=True,
                     download_name=f"results_{batch_id[:8]}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ===========================================================
# 路由 ———— 多职位矩阵模式
# ===========================================================
MULTI_HTML = """
<!doctype html><html lang="zh"><head>
<meta charset="utf-8"><title>多职位矩阵 · Alsos</title>
<link href="/static/style.css" rel="stylesheet">
<style>.card{padding:16px;border:1px solid #eee;border-radius:12px;margin:16px 0}</style>
</head><body>
<header class="hero small"><div class="brand"><div class="logo">AT</div><h1>ALSOS TALENT</h1></div></header>
<div class="container">
  <h2>多职位 × 批量简历 · 矩阵分析</h2>
  <form method="post" action="/start_multi" enctype="multipart/form-data" class="card">
    <h3>1) 上传职位 JD（可多选）</h3>
    <input type="file" name="jds" multiple accept=".pdf,.docx,.txt,.doc"/>

    <h3 style="margin-top:12px">2) 上传简历（可多选）</h3>
    <input type="file" name="resumes" multiple accept=".pdf,.docx,.txt,.doc"/>

    <h3 style="margin-top:12px">3) 报告语言</h3>
    <select name="report_lang">
      <option value="zh">中文</option>
      <option value="en">English</option>
      <option value="bi">中英双语</option>
    </select>

    <h3 style="margin-top:12px">4) 全局补充说明（可选，识别 MUST）</h3>
    <textarea name="notes" rows="5" style="width:100%" placeholder="示例：&#10;必须：英语流利&#10;[MUST] 具备药物警戒经验"></textarea>

    <div style="margin-top:16px">
      <button class="btn" type="submit">开始分析</button>
      <a class="btn" href="/">返回单职位模式</a>
    </div>
  </form>
</div>
</body></html>
"""

@app.route("/multi", methods=["GET"])
def multi_index():
    return render_template_string(MULTI_HTML)

@app.route("/start_multi", methods=["POST"])
def start_multi():
    report_lang = (request.form.get("report_lang") or "zh").lower()
    if report_lang not in ("zh","en","bi"): report_lang = "zh"

    jd_items = []
    for f in request.files.getlist("jds"):
        if not f or not allowed(f.filename): continue
        txt = parse_text_from_blob(f.filename, f.read())
        if txt: jd_items.append({"name": f.filename, "text": txt})
    if not jd_items: return "请至少上传 1 个 JD（pdf/docx/txt/doc）", 400

    resumes = []
    for f in request.files.getlist("resumes"):
        if not f or not allowed(f.filename): continue
        resumes.append({"filename": f.filename, "blob": f.read()})
    if not resumes: return "请至少上传 1 份简历", 400
    if len(jd_items) * len(resumes) > MAX_FILES * MAX_FILES:
        return "任务规模过大，请分批分析。", 400

    notes = request.form.get("notes","").strip()
    must_list = extract_must_from_notes(notes)
    batch_id = str(uuid.uuid4())
    BATCHES[batch_id] = {
        "mode": "multi",
        "jd_list": jd_items,
        "notes": notes,
        "must": must_list,
        "report_lang": report_lang,
        "pairs": [],
        "done": False,
        "csv": None,
        "excel": None,
        "stats": {"total": len(jd_items)*len(resumes), "done": 0}
    }

    def run_multi():
        def work_pair(res_item, jd_item):
            text = parse_text_from_blob(res_item["filename"], res_item["blob"])
            if not text:
                return {"resume": res_item["filename"], "jd": jd_item["name"], "ok": False, "score": 0,
                        "error":"解析失败或空文档"}
            sys_p, user_p = build_prompt(jd_item["text"], notes, must_list, text, lang=report_lang)
            ok, data = call_llm(sys_p, user_p)
            if ok and (not data.get("estimated_age") or data["estimated_age"] == "不详"):
                data["estimated_age"] = estimate_age_from_text(text)
            if not ok:
                return {"resume": res_item["filename"], "jd": jd_item["name"], "ok": False, "score": 0,
                        "error": data.get("error","LLM失败")}
            return {"resume": res_item["filename"], "jd": jd_item["name"], "ok": True,
                    "score": data.get("overall",0), "data": data}

        results = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futs = [ex.submit(work_pair, r, jd) for r in resumes for jd in jd_items]
            for fu in as_completed(futs):
                out = fu.result(); results.append(out)
                BATCHES[batch_id]["pairs"].append(out)
                BATCHES[batch_id]["stats"]["done"] += 1

        jd_names = [j["name"] for j in jd_items]
        resume_names = sorted({x["resume"] for x in results})
        scores = {rn: {jn: 0 for jn in jd_names} for rn in resume_names}
        details_by_jd = {jn: [] for jn in jd_names}
        for x in results:
            rn, jn = x["resume"], x["jd"]
            sc = x.get("score",0) or 0
            scores[rn][jn] = sc
            if x.get("ok"):
                d = x["data"]; ss = d.get("subscores",{})
                details_by_jd[jn].append({
                    "resume": rn,
                    "name": d.get("name","未知"),
                    "overall": sc,
                    "ss": [ss.get("skills",0), ss.get("education",0), ss.get("experience",0), ss.get("language",0), ss.get("stability",0)],
                    "summary": (d.get("summary","") or ""),
                    "fit": (d.get("fit_analysis","") or "").replace("\n"," "),
                })
            else:
                details_by_jd[jn].append({
                    "resume": rn,"name":"未知","overall":0,"ss":[0,0,0,0,0],"summary":"","fit":f"ERR: {x.get('error','')}"
                })

        # CSV（矩阵）
        csv_buf = io.StringIO(); w = csv.writer(csv_buf)
        w.writerow(["候选人/Resume"] + jd_names)
        for rn in resume_names:
            w.writerow([rn] + [scores[rn][jn] for jn in jd_names])
        BATCHES[batch_id]["csv"] = io.BytesIO(csv_buf.getvalue().encode("utf-8"))

        # Excel：Matrix + 每 JD sheet（含雷达图）
        lang = report_lang
        wb = Workbook()
        ws = wb.active; ws.title = "Matrix"
        ws.append(["候选人" if lang=="zh" else ("Resume" if lang=="en" else "候选人/Resume")] + jd_names)
        for rn in resume_names:
            ws.append([rn] + [scores[rn][jn] for jn in jd_names])

        for jn in jd_names:
            wsj = wb.create_sheet(title=(jn[:25] if len(jn)>25 else jn))
            wsj.append(HEADERS[lang])
            sorted_lst = sorted(details_by_jd[jn], key=lambda d: d["overall"], reverse=True)
            for d in sorted_lst:
                ss = d["ss"]
                wsj.append([
                    d["resume"], d["name"], "", "", d["summary"], "",
                    d["overall"], d["fit"], "", "",
                    ss[0], ss[1], ss[2], ss[3], ss[4]
                ])
            # 取 Top3 画雷达
            tops = []
            for d in sorted_lst[:3]:
                nm = d["name"] or d["resume"]
                tops.append([nm, *d["ss"]])
            if tops:
                add_radar_for_top(wsj, start_row=2+len(sorted_lst)+2, start_col=2, top_rows=tops)

        excel_buf = io.BytesIO(); wb.save(excel_buf); excel_buf.seek(0)
        BATCHES[batch_id]["excel"] = excel_buf
        BATCHES[batch_id]["done"] = True

    threading.Thread(target=run_multi, daemon=True).start()
    return redirect(url_for("multi_results", batch_id=batch_id))

MULTI_RESULTS_HTML = """
<!doctype html><html lang="zh"><head>
<meta charset="utf-8"><title>矩阵分析结果 · Alsos</title>
<link href="/static/style.css" rel="stylesheet">
<style>
  table{width:100%;border-collapse:collapse}
  th,td{border:1px solid #eee;padding:8px;vertical-align:top}
  .muted{color:#6b7280}
  .ok{color:#16a34a;font-weight:700}
  .err{color:#dc2626;font-weight:700}
</style>
</head><body>
<header class="hero small"><div class="brand"><div class="logo">AT</div><h1>ALSOS TALENT</h1></div></header>
<div class="container">
  <h2>多职位 · 实时进度</h2>
  <div class="actions">
    <a class="btn" href="/download_excel_multi/{{ bid }}" target="_blank">⬇ 下载Excel（矩阵+各职位）</a>
    <a class="btn" href="/download_csv_multi/{{ bid }}" target="_blank">⬇ 下载CSV（矩阵）</a>
    <a class="btn" href="/multi">返回</a>
  </div>
  <p class="muted">完成 <span id="done">0</span> / <span id="total">{{ total }}</span> …</p>
  <table id="tb">
    <thead><tr><th>简历/Resume</th><th>职位/JD</th><th>分数/Score</th><th>状态/Status</th></tr></thead>
    <tbody></tbody>
  </table>
</div>
<script>
  const tb = document.querySelector('#tb tbody');
  function esc(s){return (s||'').toString().replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
  let done = 0; const total = {{ total }};
  const es = new EventSource("/events_multi/{{ bid }}");
  es.onmessage = (e)=>{
    const x = JSON.parse(e.data);
    const tr = document.createElement('tr');
    tr.innerHTML = `<td>${esc(x.resume)}</td>
                    <td>${esc(x.jd)}</td>
                    <td>${esc(x.score)}</td>
                    <td>${x.ok?'<span class="ok">OK</span>':'<span class="err">ERR</span>'} ${x.error?('<div class="muted">'+esc(x.error)+'</div>'):''}</td>`;
    tb.appendChild(tr);
    done += 1; document.querySelector('#done').textContent = done;
  };
  es.addEventListener('done', ()=>{ document.querySelector('#done').textContent = total; });
</script>
</body></html>
"""

@app.route("/multi_results/<batch_id>")
def multi_results(batch_id):
    b = BATCHES.get(batch_id)
    if not b or b.get("mode") != "multi": return "Batch not found", 404
    return render_template_string(MULTI_RESULTS_HTML, bid=batch_id, total=b["stats"]["total"])

@app.route("/events_multi/<batch_id>")
def events_multi(batch_id):
    def gen():
        sent = 0
        while True:
            b = BATCHES.get(batch_id)
            if not b: break
            arr = b.get("pairs", [])
            while sent < len(arr):
                yield f"data: {json.dumps(arr[sent], ensure_ascii=False)}\n\n"; sent += 1
            if b["done"]:
                yield "event: done\ndata: end\n\n"; break
            time.sleep(0.8)
    return Response(gen(), mimetype="text/event-stream")

@app.route("/download_csv_multi/<batch_id>")
def download_csv_multi(batch_id):
    b = BATCHES.get(batch_id)
    if not b or not b.get("csv"): return "CSV not ready", 404
    b["csv"].seek(0)
    return send_file(b["csv"], as_attachment=True,
                     download_name=f"matrix_{batch_id[:8]}.csv",
                     mimetype="text/csv")

@app.route("/download_excel_multi/<batch_id>")
def download_excel_multi(batch_id):
    b = BATCHES.get(batch_id)
    if not b or not b.get("excel"): return "Excel not ready", 404
    b["excel"].seek(0)
    return send_file(b["excel"], as_attachment=True,
                     download_name=f"matrix_{batch_id[:8]}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ===========================================================
# 诊断
# ===========================================================
@app.route("/diag")
def diag():
    info = {
        "has_key": bool(DEEPSEEK_API_KEY),
        "base_url": OPENAI_BASE_URL,
        "model": LLM_MODEL_CHEAP,
        "max_workers": MAX_WORKERS,
    }
    try:
        _ = client.chat.completions.create(
            model=LLM_MODEL_CHEAP,
            messages=[{"role":"user","content":"ping"}],
            max_tokens=5,
            response_format={"type":"json_object"},
        )
        info["test"] = "ok"
    except Exception as e:
        info["test"] = f"error: {e}"
    return jsonify(info), 200

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
